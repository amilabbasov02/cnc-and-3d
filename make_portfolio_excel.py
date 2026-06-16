# -*- coding: utf-8 -*-
"""KAYZEN portfel — birləşmiş 48 aylıq maliyyə modeli (4 məhsul, axan komanda)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import portfolio as P, model as FR

INK="0D0E10"; AMBER="FFB300"; DARK="15171A"; LIGHT="EEF1F4"
GREEN="C9F2D8"; RED="F8D4D4"; BLUE="D6E6FF"; CYAN="D6F5F2"
PCOL={"FEEDRATE":"FFE4A8","QUOTEFLOW":"CDE7FF","FORMCHECK":"D9F2DD","CONFIGFLOW":"EAD9F7"}
def fill(h): return PatternFill("solid", fgColor=h)
thin=Side(style="thin", color="C9CDD2"); border=Border(thin,thin,thin,thin)
wrap=Alignment(wrap_text=True, vertical="top")
center=Alignment(horizontal="center", vertical="center", wrap_text=True)
def title_row(ws,row,text,span,bg=AMBER,fg=INK,size=13):
    c=ws.cell(row=row,column=1,value=text); c.font=Font(bold=True,size=size,color=fg)
    c.fill=fill(bg); c.alignment=Alignment(vertical="center")
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span); ws.row_dimensions[row].height=26
def header(ws,row,hs,bg=DARK,fg="FFFFFF"):
    for j,h in enumerate(hs,1):
        c=ws.cell(row=row,column=j,value=h); c.font=Font(bold=True,color=fg,size=9)
        c.fill=fill(bg); c.alignment=center; c.border=border

rows,meta = P.compute()
rows0,_ = P.compute(capital=0); trough0=min(r['cum'] for r in rows0)
def at(m): return rows[m-1]
wb=openpyxl.Workbook()

# ============ XÜLASƏ ============
ws=wb.active; ws.title="Xülasə (Investor)"; ws.sheet_view.showGridLines=False
for col,w in zip("ABCDEF",[40,15,15,15,15,4]): ws.column_dimensions[col].width=w
title_row(ws,1,"KAYZEN — PORTFEL İNVESTOR XÜLASƏSİ",6)
ws.cell(row=2,column=1,value="Tək şirkət · paylaşılan texnoloji nüvə · ardıcıl çıxan 4 məhsul · axan komanda (boş qalma/işdən çıxarma yox)").font=Font(italic=True,size=10)
ws.merge_cells("A2:F2")
title_row(ws,4,"İNVESTİSİYA TƏLƏBİ",6,bg=CYAN)
ask=[("Tələb olunan seed (bufer ilə)", round(-trough0*1.25,-4), "Kapitalsız ən dərin kassa × 1.25"),
 ("Modeldə qoyulan kapital", meta['capital'], "portfolio.py A['capital']"),
 ("Ən aşağı kassa nöqtəsi (runway dibi)", min(r['cum'] for r in rows), "müsbət qalır"),
 ("Kapitalsız funding need (dib)", trough0, "real ehtiyac")]
r=5
for name,val,note in ask:
    ws.cell(row=r,column=1,value=name).font=Font(bold=True)
    c=ws.cell(row=r,column=2,value=val); c.number_format='#,##0'; c.font=Font(bold=True,color="C98A06"); c.fill=fill(AMBER)
    ws.cell(row=r,column=4,value=note).font=Font(italic=True,size=9); ws.merge_cells(start_row=r,start_column=4,end_row=r,end_column=6)
    for cc in (1,2): ws.cell(row=r,column=cc).border=border
    r+=1
r+=1
title_row(ws,r,"HARDA OLACAĞIQ — 18 / 36 / 48 AY",6,bg=CYAN); r+=1
header(ws,r,["Göstərici","18 ay","36 ay","48 ay",""]); r+=1
def snap(name,fn,fmt='#,##0',bold=False):
    global r
    ws.cell(row=r,column=1,value=name).font=Font(bold=bold)
    for j,m in zip((2,3,4),(18,36,48)):
        c=ws.cell(row=r,column=j,value=fn(at(m))); c.number_format=fmt; c.alignment=center
        if bold: c.font=Font(bold=True)
    for cc in range(1,5): ws.cell(row=r,column=cc).border=border
    r+=1
snap("Aylıq gəlir (cəmi) $",lambda x:x['rev'],bold=True)
snap("İllik run-rate ARR ≈ gəlir×12 $",lambda x:x['rev']*12,bold=True)
snap("FEEDRATE gəlir/ay $",lambda x:x['prod']['FEEDRATE'])
snap("QUOTEFLOW gəlir/ay $",lambda x:x['prod']['QUOTEFLOW'])
snap("FORMCHECK gəlir/ay $",lambda x:x['prod']['FORMCHECK'])
snap("CONFIGFLOW gəlir/ay $",lambda x:x['prod']['CONFIGFLOW'])
snap("Aylıq NET $",lambda x:x['net'],bold=True)
snap("Yığılmış kassa $",lambda x:x['cum'])
snap("Canlı məhsul sayı",lambda x:x['live'],'0')
snap("Komanda (nəfər)",lambda x:x['head'],'0')
r+=1
ws.cell(row=r,column=1,value=f"Kapital səmərəliliyi: 48 ayda {at(48)['head']} nəfərlə 4 məhsul — 4 ayrı şirkət ~72 nəfər istəyərdi. Nüvə ~70% təkrar işlədilir.").font=Font(italic=True,bold=True,color="2E7D32")
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); r+=1
ws.cell(row=r,column=1,value="Valyuta USD (AZN üçün ×1.70).").font=Font(italic=True,size=9)
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)

# ============ MƏHSUL TIMELINE ============
wt=wb.create_sheet("Məhsul Timeline"); wt.sheet_view.showGridLines=False
title_row(wt,1,"MƏHSUL TIMELINE — BUILD (B) / LAUNCH→CANLI (█) · komanda axır, boş qalma yox",50)
wt.column_dimensions["A"].width=16
for c in range(2,50): wt.column_dimensions[get_column_letter(c)].width=2.6
# month header
wt.cell(row=2,column=1,value="Məhsul \\ Ay").font=Font(bold=True,size=9)
for m in range(1,49):
    cc=wt.cell(row=2,column=1+m,value=m); cc.font=Font(size=8); cc.alignment=center
prods=[("FEEDRATE",1,FR.A["launch_month"]),("QUOTEFLOW",9,12),("FORMCHECK",17,20),("CONFIGFLOW",25,28)]
r=3
for name,bstart,launch in prods:
    wt.cell(row=r,column=1,value=name).font=Font(bold=True)
    for m in range(1,49):
        cc=wt.cell(row=r,column=1+m)
        if m<bstart: continue
        if m<launch:
            cc.value="B"; cc.fill=fill(AMBER); cc.alignment=center; cc.font=Font(size=8,bold=True)
        else:
            cc.fill=fill(PCOL[name]); cc.value="█"; cc.alignment=center; cc.font=Font(size=8,color=PCOL[name])
    r+=1
r+=1
wt.cell(row=r,column=1,value="B = Build (quruluş)  ·  rəngli = Launch sonrası canlı/gəlir").font=Font(italic=True,size=9)
wt.merge_cells(start_row=r,start_column=1,end_row=r,end_column=20); r+=2
flows=["Ay 6: FEEDRATE launch → komandanın ~50%-i QUOTEFLOW build-ə keçir (boş qalma yox)",
 "Ay 12: QUOTEFLOW launch → nüvə komanda FORMCHECK-ə yönəlir",
 "Ay 20: FORMCHECK launch → CONFIGFLOW hazırlığı",
 "Ay 28: CONFIGFLOW launch → 4 məhsul paralel canlı, komanda miqyas + dəstək"]
for f in flows:
    wt.cell(row=r,column=1,value="• "+f).alignment=wrap; wt.merge_cells(start_row=r,start_column=1,end_row=r,end_column=30); wt.row_dimensions[r].height=18; r+=1

# ============ KOMANDA AXINI ============
wk=wb.create_sheet("Komanda Axını"); wk.sheet_view.showGridLines=False
for col,w in zip("ABCDE",[6,40,12,14,16]): wk.column_dimensions[col].width=w
title_row(wk,1,"PAYLAŞILAN KOMANDA — kim, nə vaxt, hansı məhsul/qrup",5)
header(wk,2,["#","Rol","Başlama","Aylıq brüt $","Məhsul / qrup"])
r=3
for i,(role,sm,sal,heavy,grp) in enumerate(P.ROLES,1):
    wk.cell(row=r,column=1,value=i).alignment=center
    wk.cell(row=r,column=2,value=role)
    wk.cell(row=r,column=3,value=f"Ay {sm}").alignment=center
    c=wk.cell(row=r,column=4,value=sal); c.number_format='#,##0'; c.alignment=center
    g=wk.cell(row=r,column=5,value=grp); g.alignment=center
    if grp in PCOL: g.fill=fill(PCOL[grp])
    for cc in range(1,6): wk.cell(row=r,column=cc).border=border
    r+=1
wk.cell(row=r,column=2,value="Tam komanda aylıq brüt (hamısı):").font=Font(bold=True)
c=wk.cell(row=r,column=4,value=f"=SUM(D3:D{r-1})"); c.number_format='#,##0'; c.font=Font(bold=True); c.fill=fill(GREEN)

# ============ 48 AY BİRLƏŞMİŞ MALİYYƏ ============
wm=wb.create_sheet("48 Ay Birləşmiş"); wm.sheet_view.showGridLines=False
cols=["Ay","Tarix","FEEDRATE $","QUOTEFLOW $","FORMCHECK $","CONFIGFLOW $","ÜMUMİ GƏLİR $",
 "Canlı məhsul","Komanda","Maaş+vergi $","Hosting $","Software $","Marketinq $","Sifariş ops $",
 "Avadanlıq $","İnzibati $","Stripe $","ÜMUMİ XƏRC $","NET $","Yığılmış kassa $"]
ws_w=[6,9,11,12,12,12,13,8,8,11,9,9,11,11,10,9,9,12,11,14]
for i,w in enumerate(ws_w,1): wm.column_dimensions[get_column_letter(i)].width=w
title_row(wm,1,"48 AYLIQ BİRLƏŞMİŞ GƏLİR-XƏRC (KAYZEN portfel)",len(cols))
header(wm,2,cols); wm.freeze_panes="C3"
for r0 in rows:
    m=r0['m']; row=2+m
    vals=[m,r0['label'],r0['prod']['FEEDRATE'],r0['prod']['QUOTEFLOW'],r0['prod']['FORMCHECK'],
     r0['prod']['CONFIGFLOW'],r0['rev'],r0['live'],r0['head'],r0['sal'],r0['host'],r0['sw'],
     r0['mkt'],r0['ops'],r0['eq'],r0['admin'],r0['stripe'],r0['cost'],r0['net'],r0['cum']]
    for j,v in enumerate(vals,1):
        c=wm.cell(row=row,column=j,value=round(v) if isinstance(v,float) else v)
        c.border=border
        if j>=3: c.number_format='#,##0'
        if j in (8,9): c.number_format='0'; c.alignment=center
    wm.cell(row=row,column=7).font=Font(bold=True); wm.cell(row=row,column=18).font=Font(bold=True)
    if m in (18,36,48):
        for j in range(1,len(cols)+1): wm.cell(row=row,column=j).fill=fill(BLUE)
wm.conditional_formatting.add("S3:T50",CellIsRule(operator='lessThan',formula=['0'],fill=fill(RED)))
tr=2+len(rows)+1
wm.cell(row=tr,column=2,value="48 AY CƏMİ").font=Font(bold=True)
for j,L in [(7,'G'),(18,'R'),(19,'S')]:
    c=wm.cell(row=tr,column=j,value=f"=SUM({L}3:{L}{2+len(rows)})"); c.number_format='#,##0'; c.font=Font(bold=True); c.fill=fill(AMBER)

# ============ İLLİK XÜLASƏ ============
wd=wb.create_sheet("İllik Xülasə"); wd.sheet_view.showGridLines=False
for col,w in zip("ABCDEF",[28,14,14,14,14,15]): wd.column_dimensions[col].width=w
title_row(wd,1,"İLLİK XÜLASƏ (USD)",6)
header(wd,2,["Göstərici","İl 1","İl 2","İl 3","İl 4","4 İL CƏMİ"])
def yr_sum(key,g):
    lo,hi={1:(1,12),2:(13,24),3:(25,36),4:(37,48)}[g]
    return sum(key(rows[m-1]) for m in range(lo,hi+1))
lines=[("FEEDRATE gəlir",lambda x:x['prod']['FEEDRATE']),
 ("QUOTEFLOW gəlir",lambda x:x['prod']['QUOTEFLOW']),
 ("FORMCHECK gəlir",lambda x:x['prod']['FORMCHECK']),
 ("CONFIGFLOW gəlir",lambda x:x['prod']['CONFIGFLOW']),
 ("ÜMUMİ GƏLİR",lambda x:x['rev']),
 ("ÜMUMİ XƏRC",lambda x:x['cost']),
 ("NET nəticə",lambda x:x['net'])]
r=3
for name,fn in lines:
    bold=name in("ÜMUMİ GƏLİR","ÜMUMİ XƏRC","NET nəticə")
    wd.cell(row=r,column=1,value=name).font=Font(bold=bold)
    tot=0
    for j,g in zip((2,3,4,5),(1,2,3,4)):
        v=yr_sum(fn,g); tot+=v
        c=wd.cell(row=r,column=j,value=round(v)); c.number_format='#,##0'; c.alignment=center
        if bold: c.font=Font(bold=True)
    c=wd.cell(row=r,column=6,value=round(tot)); c.number_format='#,##0'; c.font=Font(bold=True); c.fill=fill(LIGHT); c.alignment=center
    for cc in range(1,7): wd.cell(row=r,column=cc).border=border
    r+=1
wd.cell(row=r,column=1,value="İl sonu yığılmış kassa").font=Font(bold=True)
for j,m in zip((2,3,4,5),(12,24,36,48)):
    c=wd.cell(row=r,column=j,value=round(at(m)['cum'])); c.number_format='#,##0'; c.font=Font(bold=True); c.fill=fill(BLUE); c.alignment=center
for cc in range(1,7): wd.cell(row=r,column=cc).border=border
r+=1
wd.cell(row=r,column=1,value="İl sonu komanda").font=Font(bold=True)
for j,m in zip((2,3,4,5),(12,24,36,48)):
    c=wd.cell(row=r,column=j,value=at(m)['head']); c.alignment=center; c.font=Font(bold=True)
for cc in range(1,7): wd.cell(row=r,column=cc).border=border

wb.save("KAYZEN-Portfel-Maliyye-48ay-v3.xlsx")
print("OK Excel saved · vərəqlər:",wb.sheetnames)
