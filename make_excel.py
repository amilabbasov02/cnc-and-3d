# -*- coding: utf-8 -*-
"""FEEDRATE — Maliyyə modeli Excel (istehsal marketplace + supplier axını)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import model as M

INK="0D0E10"; AMBER="FFB300"; DARK="15171A"; LIGHT="EEF1F4"
GREEN="C9F2D8"; RED="F8D4D4"; BLUE="D6E6FF"; CYAN="D6F5F2"
def fill(h): return PatternFill("solid", fgColor=h)
thin=Side(style="thin", color="C9CDD2"); border=Border(thin,thin,thin,thin)
wrap=Alignment(wrap_text=True, vertical="top")
center=Alignment(horizontal="center", vertical="center", wrap_text=True)

wb=openpyxl.Workbook()

def title_row(ws,row,text,span,bg=AMBER,fg=INK,size=13):
    c=ws.cell(row=row,column=1,value=text); c.font=Font(bold=True,size=size,color=fg)
    c.fill=fill(bg); c.alignment=Alignment(vertical="center")
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=span)
    ws.row_dimensions[row].height=26
def header(ws,row,hs,bg=DARK,fg="FFFFFF"):
    for j,h in enumerate(hs,1):
        c=ws.cell(row=row,column=j,value=h); c.font=Font(bold=True,color=fg,size=9)
        c.fill=fill(bg); c.alignment=center; c.border=border

# ============ SHEET: XÜLASƏ / INVESTOR (birinci) ============
ws0=wb.active; ws0.title="Xülasə (Investor)"; ws0.sheet_view.showGridLines=False
for col,w in zip("ABCDEF",[40,16,16,16,16,4]): ws0.column_dimensions[col].width=w
title_row(ws0,1,"FEEDRATE — İNVESTOR XÜLASƏSİ",6)
ws0.cell(row=2,column=1,value="CNC frezeleme & 3D-çap üçün anında qiymət → sifariş → supplier istehsalı → çatdırılma platforması").font=Font(italic=True,size=10)
ws0.merge_cells("A2:F2")

# Investment ask block
title_row(ws0,4,"İNVESTİSİYA TƏLƏBİ",6,bg=CYAN)
DASH="'36 Aylıq Maliyyə'"; FZ="'Fərziyyələr'"
ask=[
 ("Tələb olunan investisiya (seed) — buferlə", f"=ROUND(MAX(0,{FZ}!B24-MIN({DASH}!W3:W50))*1.15,-3)", "Ən dərin kassa çuxuru + 15% bufer"),
 ("Başlanğıc kapital (modeldə qoyulan)", f"={FZ}!B24", "Fərziyyələr!B24-də dəyiş"),
 ("Ən aşağı kassa nöqtəsi (runway dibi)", f"=MIN({DASH}!W3:W50)", "Bu müsbət qalmalıdır"),
 ("Breakeven ayı (ilk müsbət NET)", f"=MATCH(TRUE,INDEX({DASH}!V3:V50>0,0),0)", "Ay nömrəsi"),
]
r=5
for name,f,note in ask:
    ws0.cell(row=r,column=1,value=name).font=Font(bold=True)
    c=ws0.cell(row=r,column=2,value=f); c.number_format='#,##0'; c.font=Font(bold=True,color="C98A06"); c.fill=fill(AMBER)
    ws0.cell(row=r,column=4,value=note).font=Font(italic=True,size=9); ws0.merge_cells(start_row=r,start_column=4,end_row=r,end_column=5)
    for cc in range(1,3): ws0.cell(row=r,column=cc).border=border
    r+=1

# Where we'll be: snapshots
r+=1
title_row(ws0,r,"HARDA OLACAĞIQ — 18 və 36 AY",6,bg=CYAN); r+=1
header(ws0,r,["Göstərici","18 ay","24 ay","36 ay","48 ay"]); r+=1
# month -> excel row in dashboard: 2+m
def MR(m): return 2+m
snap=[
 ("Aylıq gəlir (MRR-ə bənzər) $","K"),
 ("İllik run-rate (ARR ≈ gəlir×12) $","K*12"),
 ("Aylıq sifariş sayı","D"),
 ("İstehsal GMV (aylıq) $","E"),
 ("Ödənişli abunəçi","G"),
 ("Aylıq NET $","V"),
 ("Yığılmış kassa $","W"),
 ("Komanda (nəfər)","L"),
]
for name,col in snap:
    ws0.cell(row=r,column=1,value=name).font=Font(bold=("NET" in name or "gəlir" in name))
    for j,m in enumerate([18,24,36,48],2):
        if col=="K*12":
            f=f"={DASH}!K{MR(m)}*12"
        else:
            f=f"={DASH}!{col}{MR(m)}"
        c=ws0.cell(row=r,column=j,value=f)
        c.number_format='0' if col=="L" else '#,##0'; c.alignment=center
        if "NET" in name or "ARR" in name.upper(): c.font=Font(bold=True)
    for cc in range(1,6): ws0.cell(row=r,column=cc).border=border
    r+=1
r+=1
ws0.cell(row=r,column=1,value="İstifadə yönü: komanda ~%55 · məhsul/infra ~%15 · marketinq ~%20 · əməliyyat/inzibati ~%10").font=Font(italic=True)
ws0.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); r+=1
ws0.cell(row=r,column=1,value="Valyuta USD (AZN üçün ×1.70). Bütün rəqəmlər «Fərziyyələr»dən avtomatik hesablanır.").font=Font(italic=True,size=9)
ws0.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)

# ============ SHEET: NECƏ İŞLƏYİR (axın) ============
wf=wb.create_sheet("Necə İşləyir (axın)"); wf.sheet_view.showGridLines=False
for col,w in zip("ABCD",[6,30,70,30]): wf.column_dimensions[col].width=w
title_row(wf,1,"MƏHSUL AXINI — QİYMƏTDƏN ÇATDIRILMAYA",4)
header(wf,2,["Addım","Mərhələ","Nə baş verir","Bizim gəlir / rol"])
flow=[
 ("1","Model","İstifadəçi: (a) kitabxanadan hazır model seçir, (b) brauzer CAD-da qurur, və ya (c) öz STL/STEP faylını yükləyir","Abunə (CAD/alət istifadəsi)"),
 ("2","Anında qiymət","Material, ölçü, miqdar, emal vaxtı əsasında saniyələrlə qiymət","—"),
 ("3","Sifariş + ödəniş","İstifadəçi sifariş verir və tam məbləği ödəyir (Stripe)","Ödənişi biz toplayırıq"),
 ("4","Supplier marşrutlaşdırma","Sifariş ən uyğun supplier emalatxanasına (CNC/3D-print) yönləndirilir","Supplier şəbəkəsi"),
 ("5","İstehsal","Supplier hissəni hazırlayır; biz statusu izləyirik","Order ops koordinasiya"),
 ("6","Keyfiyyət + çatdırılma","Keyfiyyət yoxlanışı → çatdırılma müştəriyə","Keyfiyyət/logistika"),
 ("7","Hesablaşma","Supplier-ə payout (Stripe Connect); biz marjanı saxlayırıq","İSTEHSAL MARJASI ~20%"),
 ("8","Dizayn satışı (opsional)","Dizayner öz modelini listing kimi satır","DİZAYN KOMİSSİYA ~8%"),
]
r=3
for n,st,what,rev in flow:
    wf.cell(row=r,column=1,value=n).alignment=center; wf.cell(row=r,column=1).font=Font(bold=True,color="C98A06")
    wf.cell(row=r,column=2,value=st).font=Font(bold=True); wf.cell(row=r,column=2).fill=fill(LIGHT)
    wf.cell(row=r,column=3,value=what).alignment=wrap
    wf.cell(row=r,column=4,value=rev).alignment=wrap
    for c in range(1,5): wf.cell(row=r,column=c).border=border
    wf.row_dimensions[r].height=40; r+=1
r+=1
wf.cell(row=r,column=2,value="3 GƏLİR MƏNBƏYİ:").font=Font(bold=True,size=11)
wf.cell(row=r+1,column=2,value="1) İstehsal marjası (əsas): sifariş GMV × ~20%")
wf.cell(row=r+2,column=2,value="2) Abunə (SaaS): CAD + qiymət alətləri, $9–$99/ay")
wf.cell(row=r+3,column=2,value="3) Dizayn marketplace komissiya: ~8%")

# ============ SHEET: KOMANDA & ROLLAR ============
wt=wb.create_sheet("Komanda & Rollar"); wt.sheet_view.showGridLines=False
for col,w in zip("ABCDEF",[4,28,12,44,12,11]): wt.column_dimensions[col].width=w
title_row(wt,1,"KOMANDA, ROLLAR, MAAŞLAR (aylıq brüt USD) — kim nə edir",6)
header(wt,2,["#","Rol / Vəzifə","Başlama","Əsas məsuliyyət","Maaş $","Tip"])
r=3
for i,(role,sm,sal,resp,tp,heavy) in enumerate(M.ROLES,1):
    wt.cell(row=r,column=1,value=i).alignment=center
    wt.cell(row=r,column=2,value=role).font=Font(bold=True)
    wt.cell(row=r,column=3,value=f"Ay {sm}").alignment=center
    wt.cell(row=r,column=4,value=resp).alignment=wrap
    c=wt.cell(row=r,column=5,value=sal); c.number_format='#,##0'; c.alignment=center
    wt.cell(row=r,column=6,value=tp).alignment=center
    for cc in range(1,7): wt.cell(row=r,column=cc).border=border
    wt.row_dimensions[r].height=30; r+=1
wt.cell(row=r,column=4,value="Tam komanda aylıq brüt (hamısı işə düşəndə):").font=Font(bold=True)
c=wt.cell(row=r,column=5,value=f"=SUM(E3:E{r-1})"); c.number_format='#,##0'; c.font=Font(bold=True); c.fill=fill(GREEN)
r+=2
wt.cell(row=r,column=2,value="Qeyd: brüt USD. Faktiki işəgötürən xərci = brüt × 1.25 (vergi/SSF, Fərziyyələr!B17).").font=Font(italic=True)
wt.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)

# ============ SHEET: AYLIQ İŞ PLANI ============
wp=wb.create_sheet("Aylıq İş Planı"); wp.sheet_view.showGridLines=False
for col,w in zip("ABCD",[14,28,62,28]): wp.column_dimensions[col].width=w
title_row(wp,1,"FAZALI İŞ PLANI — KİM NƏ EDİR / NƏ ÇATDIRILIR",4)
header(wp,2,["Dövr","Faza","Kim nə edir","Çatdırılma"])
plan=[
 ("Ay 1–3","Təməl","CEO: kapital, hüquqi qeydiyyat, ilk supplier danışıqları. CTO: arxitektura + qiymət mühərriki + 3D. Frontend: əsas UI + quote.","Auth + DB + limit + işlək quote/3D"),
 ("Ay 3–5","Kitabxana + ödəniş","CTO/Frontend: model qaleriyası + parametrik ölçü. Backend: Stripe sifariş/ödəniş. UI/UX: checkout axını.","Lisenziyalı model + sifariş & ödəniş axını"),
 ("Ay 5–7","Supplier şəbəkəsi","Ops/Supplier Manager: ilk emalatxanaları cəlb, qiymət/keyfiyyət razılaşması, marşrutlaşdırma. Backend: Stripe Connect payout.","İlk supplier-lərlə canlı sifariş icrası"),
 ("Ay 7–10","CAD MVP + marketinq","CTO: three-bvh-csg redaktor (export). Marketinq: SEO + launch. Frontend: redaktor UI.","Brauzer CAD v1 + ilk müştəri axını"),
 ("Ay 10–14","Marketplace + əməliyyat","Order ops/Logistika: sifariş izləmə, keyfiyyət, çatdırılma. 3D kontent: kitabxana. Dəstək: onboarding.","Tam sifariş→supplier→çatdırılma dövrü"),
 ("Ay 14–24","Böyümə","Marketinq: CAC/LTV, kanallar. Ops: daha çox supplier, regional əhatə. Financist: aylıq hesabat, runway.","Breakeven-ə çatma, supplier şəbəkəsi genişlənir"),
 ("Ay 24–36","Miqyas & B2B","Sales/BizDev: Business/API, korporativ. CTO: STEP export, API. Komanda: stabil əməliyyat.","B2B müştərilər, beynəlxalq genişlənmə, mənfəət"),
]
r=3
for d,ph,who,deliv in plan:
    wp.cell(row=r,column=1,value=d).font=Font(bold=True); wp.cell(row=r,column=1).fill=fill(BLUE); wp.cell(row=r,column=1).alignment=center
    wp.cell(row=r,column=2,value=ph).font=Font(bold=True); wp.cell(row=r,column=2).alignment=wrap
    wp.cell(row=r,column=3,value=who).alignment=wrap
    wp.cell(row=r,column=4,value=deliv).alignment=wrap
    for c in range(1,5): wp.cell(row=r,column=c).border=border
    wp.row_dimensions[r].height=58; r+=1

# ============ SHEET: KOMPÜTER & AVADANLIQ ============
we=wb.create_sheet("Kompüter & Avadanlıq"); we.sheet_view.showGridLines=False
for col,w in zip("ABCDE",[36,10,12,14,40]): we.column_dimensions[col].width=w
title_row(we,1,"KOMPÜTER, AVADANLIQ & PROQRAM (USD)",5)
header(we,2,["Avadanlıq","Sayı","Vahid $","Cəmi $","Qeyd"])
equip=[("Developer iş stansiyası (GPU — CAD/3D)",5,1800,"Three.js/CSG render üçün"),
 ("Ofis noutbuku (ops/marketinq/financist/dəstək)",6,900,"Qeyri-tex rollar"),
 ("Monitor + klaviatura + siçan",9,300,"Hər iş yeri"),
 ("Test cihazları (telefon/tablet)",1,600,"Mobil/responsiv test"),
 ("Şəbəkə / NAS / backup",1,700,"Yedəkləmə")]
r=3; es=r
for n,q,u,note in equip:
    we.cell(row=r,column=1,value=n).alignment=wrap; we.cell(row=r,column=2,value=q).alignment=center
    we.cell(row=r,column=3,value=u).number_format='#,##0'; we.cell(row=r,column=3).alignment=center
    c=we.cell(row=r,column=4,value=f"=B{r}*C{r}"); c.number_format='#,##0'; c.alignment=center
    we.cell(row=r,column=5,value=note).alignment=wrap
    for cc in range(1,6): we.cell(row=r,column=cc).border=border
    r+=1
we.cell(row=r,column=1,value="AVADANLIQ CƏMİ (birdəfəlik, fazalı)").font=Font(bold=True)
c=we.cell(row=r,column=4,value=f"=SUM(D{es}:D{r-1})"); c.number_format='#,##0'; c.font=Font(bold=True); c.fill=fill(GREEN)
r+=2
header(we,r,["Proqram / Servis (aylıq)","Sayı","Aylıq $","İllik $","Qeyd"]); r+=1; ss=r
sw=[("Hosting / Vercel / infra",1,150,"trafikə görə artır"),
 ("PostgreSQL (Neon/Supabase)",1,60,"DB"),
 ("Cloudflare R2 / S3",1,40,"fayl/model saxlama"),
 ("Stripe (Subscriptions+Connect)",1,0,"~3% faiz — modeldə ayrıca"),
 ("Google Workspace + domen",13,8,"hər nəfər/ay"),
 ("Figma + analitika + alətlər",1,120,"komanda lisenziya"),
 ("GitHub / CI / monitoring",1,80,"DevOps")]
for n,q,m_,note in sw:
    we.cell(row=r,column=1,value=n).alignment=wrap; we.cell(row=r,column=2,value=q).alignment=center
    we.cell(row=r,column=3,value=m_).number_format='#,##0'; we.cell(row=r,column=3).alignment=center
    c=we.cell(row=r,column=4,value=f"=B{r}*C{r}*12"); c.number_format='#,##0'; c.alignment=center
    we.cell(row=r,column=5,value=note).alignment=wrap
    for cc in range(1,6): we.cell(row=r,column=cc).border=border
    r+=1
we.cell(row=r,column=1,value="PROQRAM CƏMİ").font=Font(bold=True)
c=we.cell(row=r,column=3,value=f"=SUM(C{ss}:C{r-1})"); c.number_format='#,##0'; c.font=Font(bold=True); c.fill=fill(LIGHT)
c=we.cell(row=r,column=4,value=f"=SUM(D{ss}:D{r-1})"); c.number_format='#,##0'; c.font=Font(bold=True); c.fill=fill(GREEN)

# ============ SHEET: FƏRZİYYƏLƏR ============
wa=wb.create_sheet("Fərziyyələr"); wa.sheet_view.showGridLines=False
wa.column_dimensions["A"].width=46; wa.column_dimensions["B"].width=14; wa.column_dimensions["C"].width=48
title_row(wa,1,"FƏRZİYYƏLƏR — dəyişin, bütün model yenilənir (Orta ssenari)",3)
header(wa,2,["Parametr","Dəyər","İzah"])
assump=[
 ("Başlanğıc aylıq ziyarətçi",M.A["vis1"],'0',"1-ci ay"),
 ("Ziyarətçi aylıq artım",M.A["vis_g"],'0%',"hər ay"),
 ("Sifariş konversiya (% ziyarətçi)",M.A["order_conv"],'0.0%',"qiymət alıb sifariş verən"),
 ("Orta sifariş dəyəri AOV ($)",M.A["aov"],'0',"CNC/3D hissə"),
 ("İstehsal marjası (take-rate)",M.A["take"],'0%',"bizə qalan pay"),
 ("Sifariş başına ops xərci ($)",M.A["ops_per_order"],'0',"keyfiyyət/koordinasiya"),
 ("Pulsuz qeydiyyat (% ziyarətçi)",M.A["free_rate"],'0%',""),
 ("Ödənişli konversiya (% yeni pulsuz)",M.A["paid_conv"],'0%',""),
 ("Aylıq churn (ödənişli)",M.A["churn"],'0%',""),
 ("Blended ARPU ($/ay)",M.A["arpu"],'0',"paketlərin ortası"),
 ("Dizayn GMV — ay1 ($)",M.A["dgmv1"],'0',""),
 ("Dizayn GMV artım",M.A["dgmv_g"],'0%',""),
 ("Dizayn komissiya",M.A["dcomm"],'0%',""),
 ("Stripe / ödəniş faizi",M.A["stripe"],'0%',"tam həcm üzərində"),
 ("Maaş vergi/SSF əmsalı",M.A["tax"],'0%',"işəgötürən"),
 ("Hosting baza ($/ay)",M.A["host_base"],'0',""),
 ("Hosting / 10K ziyarətçi ($)",M.A["host_10k"],'0',""),
 ("Software / nəfər ($/ay)",M.A["sw_head"],'0',""),
 ("Marketinq — % gəlir",M.A["mkt_pct"],'0%',""),
 ("Marketinq minimum ($/ay)",M.A["mkt_min"],'0',""),
 ("İnzibati baza ($/ay)",M.A["admin"],'0',""),
 ("Başlanğıc kapital / investisiya ($)",M.A["capital"],'0',"seed"),
 ("Launch ayı — gəlir bu aydan başlayır",M.A["launch_month"],'0',"1-5 quruluş, gəlir YOXDUR"),
]
r=3
for name,val,fmt,izah in assump:
    wa.cell(row=r,column=1,value=name).alignment=wrap
    c=wa.cell(row=r,column=2,value=val); c.number_format=fmt; c.font=Font(bold=True); c.fill=fill(AMBER); c.alignment=center
    wa.cell(row=r,column=3,value=izah).alignment=wrap
    for cc in range(1,4): wa.cell(row=r,column=cc).border=border
    r+=1
def AR(idx): return f"{FZ}!$B${2+idx}"

# ============ SHEET: 36 AYLIQ MALİYYƏ ============
wm=wb.create_sheet("36 Aylıq Maliyyə"); wm.sheet_view.showGridLines=False
cols=["Ay","Tarix","Ziyarətçi","Sifariş","İstehsal GMV $","İstehsal gəliri $","Ödənişli ist.",
 "Abunə gəliri $","Dizayn GMV $","Dizayn kom. $","ÜMUMİ GƏLİR $","Komanda","Maaş+vergi $",
 "Hosting $","Software $","Marketinq $","Sifariş ops $","Avadanlıq $","İnzibati $","Stripe $",
 "ÜMUMİ XƏRC $","NET $","Yığılmış kassa $"]
for i,w in enumerate([6,9,10,8,12,12,9,10,11,11,12,9,11,9,9,10,10,10,9,9,12,11,14],1):
    wm.column_dimensions[get_column_letter(i)].width=w
title_row(wm,1,"36 AYLIQ GƏLİR-XƏRC (canlı düstur — Fərziyyələr dəyişdikcə yenilənir)",23)
header(wm,2,cols); wm.freeze_panes="C3"

for m in range(1,49):
    row=2+m; r=row
    wm.cell(row=r,column=1,value=m).alignment=center
    wm.cell(row=r,column=2,value=M.month_label(m)).alignment=center
    # C vis (launch ayına qədər 0)
    wm.cell(row=r,column=3,value=f"=IF(A{r}<{AR(23)},0,IF(A{r}={AR(23)},{AR(1)},C{r-1}*(1+{AR(2)})))")
    # D orders
    wm.cell(row=r,column=4,value=f"=C{r}*{AR(3)}")
    # E mgmv
    wm.cell(row=r,column=5,value=f"=D{r}*{AR(4)}")
    # F mrev
    wm.cell(row=r,column=6,value=f"=E{r}*{AR(5)}")
    # G paid
    if m==1: wm.cell(row=r,column=7,value=f"=C{r}*{AR(7)}*{AR(8)}")
    else: wm.cell(row=r,column=7,value=f"=G{r-1}*(1-{AR(9)})+C{r}*{AR(7)}*{AR(8)}")
    # H sub
    wm.cell(row=r,column=8,value=f"=G{r}*{AR(10)}")
    # I dgmv (launch ayına qədər 0)
    wm.cell(row=r,column=9,value=f"=IF(A{r}<{AR(23)},0,IF(A{r}={AR(23)},{AR(11)},I{r-1}*(1+{AR(12)})))")
    # J dcom
    wm.cell(row=r,column=10,value=f"=I{r}*{AR(13)}")
    # K rev
    wm.cell(row=r,column=11,value=f"=F{r}+H{r}+J{r}")
    # L head
    wm.cell(row=r,column=12,value=M.headcount(m)).alignment=center
    # M sal
    wm.cell(row=r,column=13,value=f"={M.gross(m)}*(1+{AR(15)})")
    # N host
    wm.cell(row=r,column=14,value=f"={AR(16)}+C{r}/10000*{AR(17)}")
    # O sw
    wm.cell(row=r,column=15,value=f"=L{r}*{AR(18)}")
    # P mkt
    wm.cell(row=r,column=16,value=f"=MAX({AR(20)},K{r}*{AR(19)})")
    # Q ops
    wm.cell(row=r,column=17,value=f"=D{r}*{AR(6)}")
    # R equip
    wm.cell(row=r,column=18,value=M.equipment(m))
    # S admin
    extra=(300 if m>=13 else 0)+(400 if m>=25 else 0)
    wm.cell(row=r,column=19,value=f"={AR(21)}+{extra}")
    # T stripe
    wm.cell(row=r,column=20,value=f"=(E{r}+H{r}+I{r})*{AR(14)}")
    # U cost
    wm.cell(row=r,column=21,value=f"=M{r}+N{r}+O{r}+P{r}+Q{r}+R{r}+S{r}+T{r}")
    # V net
    wm.cell(row=r,column=22,value=f"=K{r}-U{r}")
    # W cum
    wm.cell(row=r,column=23,value=f"={AR(22)}+V{r}" if m==1 else f"=W{r-1}+V{r}")
    for c in range(3,24): wm.cell(row=r,column=c).number_format='#,##0'
    wm.cell(row=r,column=12).number_format='0'
    for c in range(1,24): wm.cell(row=r,column=c).border=border
    wm.cell(row=r,column=11).font=Font(bold=True); wm.cell(row=r,column=21).font=Font(bold=True)
    # highlight month 18 & 36
    if m in (18,36,48):
        for c in range(1,24): wm.cell(row=r,column=c).fill=fill(BLUE)

wm.conditional_formatting.add("V3:W50",CellIsRule(operator='lessThan',formula=['0'],fill=fill(RED)))
wm.conditional_formatting.add("V3:V50",CellIsRule(operator='greaterThanOrEqual',formula=['0'],fill=fill(GREEN)))
# totals
tr=51
wm.cell(row=tr,column=2,value="48 AY CƏMİ").font=Font(bold=True)
for col,L in [(11,'K'),(21,'U'),(22,'V')]:
    c=wm.cell(row=tr,column=col,value=f"=SUM({L}3:{L}50)"); c.number_format='#,##0'; c.font=Font(bold=True); c.fill=fill(AMBER)
for c in range(1,24): wm.cell(row=tr,column=c).border=border

# ============ SHEET: İLLİK XÜLASƏ ============
wd=wb.create_sheet("İllik Xülasə"); wd.sheet_view.showGridLines=False
for col,w in zip("ABCDEF",[32,14,14,14,14,15]): wd.column_dimensions[col].width=w
title_row(wd,1,"İLLİK XÜLASƏ (USD)",6)
header(wd,2,["Göstərici","İl 1","İl 2","İl 3","İl 4","4 İL CƏMİ"])
rows=[("İstehsal gəliri (marja)","F"),("Abunə gəliri","H"),("Dizayn komissiya","J"),
 ("ÜMUMİ GƏLİR","K"),("Maaş+vergi","M"),("Marketinq","P"),("Sifariş ops","Q"),
 ("ÜMUMİ XƏRC","U"),("NET nəticə","V")]
ranges=[(3,14),(15,26),(27,38),(39,50)]; r=3
for name,col in rows:
    bold=name in("ÜMUMİ GƏLİR","ÜMUMİ XƏRC","NET nəticə")
    wd.cell(row=r,column=1,value=name).font=Font(bold=bold)
    for j,(a,b) in enumerate(ranges,2):
        c=wd.cell(row=r,column=j,value=f"=SUM({DASH}!{col}{a}:{col}{b})"); c.number_format='#,##0'; c.alignment=center
        if bold: c.font=Font(bold=True)
    c=wd.cell(row=r,column=6,value=f"=SUM(B{r}:E{r})"); c.number_format='#,##0'; c.font=Font(bold=True); c.fill=fill(LIGHT); c.alignment=center
    for cc in range(1,7): wd.cell(row=r,column=cc).border=border
    r+=1
wd.cell(row=r,column=1,value="İl sonu yığılmış kassa").font=Font(bold=True)
for j,end in zip(range(2,6),[14,26,38,50]):
    c=wd.cell(row=r,column=j,value=f"={DASH}!W{end}"); c.number_format='#,##0'; c.font=Font(bold=True); c.fill=fill(BLUE); c.alignment=center
for cc in range(1,7): wd.cell(row=r,column=cc).border=border
r+=1
wd.cell(row=r,column=1,value="İl sonu komanda (nəfər)").font=Font(bold=True)
for j,end in zip(range(2,6),[14,26,38,50]):
    c=wd.cell(row=r,column=j,value=f"={DASH}!L{end}"); c.alignment=center; c.font=Font(bold=True)
for cc in range(1,7): wd.cell(row=r,column=cc).border=border

wb.save("FEEDRATE-Maliyye-Model-Premium.xlsx")
print("OK Excel saved")
